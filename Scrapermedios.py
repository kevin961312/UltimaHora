"""
Scrapers de medios de comunicación colombianos — noticias de última hora.
Cada función retorna lista de dicts:
  {"titulo": str, "url": str, "medio": str, "fecha": str}
run_all() ejecuta todos en paralelo y ordena de más reciente a más antigua.
"""

import re
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional, List, Dict

import requests
from bs4 import BeautifulSoup
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

_MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

_WORD_NUMS = {
    "un": 1, "una": 1, "uno": 1, "dos": 2, "tres": 3, "cuatro": 4,
    "cinco": 5, "seis": 6, "siete": 7, "ocho": 8, "nueve": 9, "diez": 10,
    "once": 11, "doce": 12, "quince": 15, "veinte": 20, "treinta": 30,
    "cuarenta": 40, "cincuenta": 50,
}


# ─────────────────────────────────────────────────────────────────────────────
# Utilidades HTTP
# ─────────────────────────────────────────────────────────────────────────────

def _get(url: str, verify=True, timeout=15) -> Optional[requests.Response]:
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, verify=verify,
                         allow_redirects=True)
        r.raise_for_status()
        return r
    except Exception as e:
        print(f"  [!] {url[:70]} → {e}")
        return None


def _soup(url: str, verify=True) -> Optional[BeautifulSoup]:
    r = _get(url, verify=verify)
    return BeautifulSoup(r.content, "html.parser") if r else None


def _abs_url(href: str, base: str) -> str:
    if not href:
        return ""
    href = href.strip()
    if href.startswith("http"):
        return href
    if href.startswith("//"):
        return "https:" + href
    return base.rstrip("/") + "/" + href.lstrip("/")


# ─────────────────────────────────────────────────────────────────────────────
# Parseo de fechas y tiempos
# ─────────────────────────────────────────────────────────────────────────────

def _parse_relative(s: str) -> Optional[datetime]:
    """Convierte 'hace 5 minutos', 'hace 2 horas', 'hace 57 min', etc. a datetime."""
    now = datetime.utcnow() - timedelta(hours=5)  # siempre COT (UTC-5)
    sl = s.lower().strip()

    if re.search(r"hace\s+(un\s+)?(momento|instante|un?\s+segundo)", sl):
        return now

    # Soporta: minuto/min, hora/h, día/dia, semana, mes + alias cortos
    m = re.search(
        r"hace\s+(\w+)\s+(min(?:uto)?|h(?:ora)?|seg(?:undo)?|d[ií]a|semana|mes)s?",
        sl
    )
    if m:
        val_str, unit = m.group(1), m.group(2)
        val = _WORD_NUMS.get(val_str)
        if val is None:
            try:
                val = int(val_str)
            except ValueError:
                return None

        if unit.startswith("seg"):
            return now - timedelta(seconds=val)
        if unit.startswith("min"):
            return now - timedelta(minutes=val)
        if unit.startswith("h"):
            return now - timedelta(hours=val)
        if unit.startswith("d"):
            return now - timedelta(days=val)
        if unit == "semana":
            return now - timedelta(weeks=val)
        if unit == "mes":
            return now - timedelta(days=val * 30)

    return None


def _parse_time_only(s: str) -> Optional[datetime]:
    """Parsea hora suelta 'HH:MM a. m.' / 'HH:MM p. m.' y la combina con fecha de hoy."""
    s = s.strip()
    m = re.match(r"(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(a\.?\s*m\.?|p\.?\s*m\.?)?$", s, re.I)
    if not m:
        return None
    hh, mm, ss = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
    meridiem = (m.group(4) or "").lower().replace(" ", "").replace(".", "")
    if meridiem == "pm" and hh < 12:
        hh += 12
    elif meridiem == "am" and hh == 12:
        hh = 0
    today = datetime.now().date()
    try:
        dt = datetime(today.year, today.month, today.day, hh, mm, ss)
        if dt > datetime.now():          # hora futura → artículo de ayer
            dt -= timedelta(days=1)
        return dt
    except ValueError:
        return None


def _parse_dt(s: str) -> Optional[datetime]:
    """String de fecha/hora → datetime. Soporta ISO, dd/mm/yyyy, texto español/inglés y relativas."""
    if not s:
        return None
    s = s.strip()

    # ISO 8601 con hora y offset opcional: 2026-05-21T17:26:32+02:00 / ...Z / ...000Z
    m = re.search(
        r"(\d{4})-(\d{2})-(\d{2})[T ](\d{2}):(\d{2})(?::(\d{2})(?:\.\d+)?)?(?P<tz>[+-]\d{2}:\d{2}|Z)?",
        s,
    )
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)),
                          int(m.group(4)), int(m.group(5)), int(m.group(6) or 0))
            tz_str = m.group("tz")
            if tz_str:
                if tz_str == "Z":
                    dt = dt - timedelta(hours=5)          # UTC → Colombia (UTC-5)
                else:
                    sign = 1 if tz_str[0] == "+" else -1
                    h, mn = map(int, tz_str[1:].split(":"))
                    dt = dt - timedelta(hours=sign * h, minutes=sign * mn) - timedelta(hours=5)
            return dt
        except ValueError:
            pass

    # ISO solo fecha: 2026-05-21
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # dd/mm/yyyy [HH:MM[:SS]]
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})(?:[T\s]+(\d{1,2}):(\d{2})(?::(\d{2}))?)?", s)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)),
                            int(m.group(4) or 0), int(m.group(5) or 0), int(m.group(6) or 0))
        except ValueError:
            pass

    # dd de MONTH [de] yyyy [HH:MM]
    m = re.search(r"(\d{1,2})\s+de\s+(\w+?)(?:\s+de)?\s+(\d{4})(?:.*?(\d{1,2}):(\d{2}))?", s, re.I)
    if m:
        month = _MESES.get(m.group(2).lower())
        if month:
            try:
                return datetime(int(m.group(3)), month, int(m.group(1)),
                                int(m.group(4) or 0), int(m.group(5) or 0))
            except ValueError:
                pass

    # RFC 2822 pubDate: Thu, 21 May 2026 18:43:36 +0000  ← ANTES de dd MONTH yyyy
    m = re.search(
        r"\w+,\s+(\d{1,2})\s+(\w+)\s+(\d{4})\s+(\d{2}):(\d{2})(?::(\d{2}))?(?:\s+([+-]\d{2}:?\d{2}|UTC|GMT|Z))?",
        s,
    )
    if m:
        month = _MESES.get(m.group(2).lower())
        if month:
            try:
                dt = datetime(int(m.group(3)), month, int(m.group(1)),
                              int(m.group(4)), int(m.group(5)), int(m.group(6) or 0))
                tz_str = (m.group(7) or "").strip()
                if tz_str in ("UTC", "GMT", "Z", "+0000", "+00:00"):
                    dt = dt - timedelta(hours=5)                  # UTC → COT
                elif re.match(r"[+-]\d{4}$", tz_str):
                    sign = 1 if tz_str[0] == "+" else -1
                    h, mn = int(tz_str[1:3]), int(tz_str[3:5])
                    dt = dt - timedelta(hours=sign * h, minutes=sign * mn) - timedelta(hours=5)
                elif re.match(r"[+-]\d{2}:\d{2}$", tz_str):
                    sign = 1 if tz_str[0] == "+" else -1
                    h, mn = map(int, tz_str[1:].split(":"))
                    dt = dt - timedelta(hours=sign * h, minutes=sign * mn) - timedelta(hours=5)
                return dt
            except ValueError:
                pass

    # dd MONTH yyyy [HH:MM]
    m = re.search(r"(\d{1,2})\s+(\w+)\s+(\d{4})(?:.*?(\d{1,2}):(\d{2}))?", s, re.I)
    if m:
        month = _MESES.get(m.group(2).lower())
        if month:
            try:
                return datetime(int(m.group(3)), month, int(m.group(1)),
                                int(m.group(4) or 0), int(m.group(5) or 0))
            except ValueError:
                pass

    # MONTH dd, yyyy [HH:MM]  (English)
    m = re.search(r"(\w+)\s+(\d{1,2}),?\s+(\d{4})(?:.*?(\d{1,2}):(\d{2}))?", s, re.I)
    if m:
        month = _MESES.get(m.group(1).lower())
        if month:
            try:
                return datetime(int(m.group(3)), month, int(m.group(2)),
                                int(m.group(4) or 0), int(m.group(5) or 0))
            except ValueError:
                pass

    # Tiempo solo: "07:22 p. m."
    dt = _parse_time_only(s)
    if dt:
        return dt

    # Relativa: "hace 5 minutos"
    return _parse_relative(s)


def _fmt_dt(dt: Optional[datetime]) -> str:
    if not dt:
        return ""
    h12 = dt.hour % 12 or 12
    meridiem = "a. m." if dt.hour < 12 else "p. m."
    return f"{dt.day:02d}/{dt.month:02d}/{dt.year} {h12:02d}:{dt.minute:02d} {meridiem}"


def _item(titulo: str, url: str, medio: str,
          fecha_str: str = "", dt: Optional[datetime] = None) -> dict:
    if dt is None and fecha_str:
        dt = _parse_dt(fecha_str)
    return {
        "titulo": titulo.strip(),
        "url": url,
        "medio": medio,
        "fecha": _fmt_dt(dt) if dt else fecha_str,
        "_dt": dt or datetime.min,
    }


def _dt_from_time_tag(tag) -> Optional[datetime]:
    if tag is None:
        return None
    for attr in ("datetime", "content"):
        val = tag.get(attr, "")
        if val:
            dt = _parse_dt(val)
            if dt:
                return dt
    return _parse_dt(tag.get_text(strip=True))


# ─────────────────────────────────────────────────────────────────────────────
# 1. El Espectador
#    URL: https://www.elespectador.com/ultimas-noticias-colombia/
#    Contenedor: BlockContainer_lastNewsSection
#    Items: div.Card-Container  →  h2 (título), p.Card-Datetime (hora), a en Card-ImageFrame (URL)
# ─────────────────────────────────────────────────────────────────────────────
def scrape_elespectador() -> List[Dict]:
    medio = "El Espectador"
    base  = "https://www.elespectador.com"
    soup  = _soup(f"{base}/ultimas-noticias-colombia/")
    if not soup:
        return []

    # Localizar el contenedor indicado
    container = None
    for tag in soup.find_all(class_=True):
        if "BlockContainer_lastNewsSection" in " ".join(tag.get("class", [])):
            container = tag
            break
    if not container:
        container = soup

    results, seen = [], set()
    for card in container.find_all("div", class_=re.compile(r"^Card$")):
        # La URL está en la imagen (Card-ImageFrame > a)
        img_frame = card.select_one(".Card-ImageFrame a[href]")
        href = img_frame.get("href", "") if img_frame else ""
        href = _abs_url(href, base)
        if not href or href in seen:
            continue
        seen.add(href)

        # Título: h2 dentro de Card-Container
        content = card.select_one(".Card-Container")
        titulo_tag = content.select_one("h2") if content else None
        titulo = titulo_tag.get_text(strip=True) if titulo_tag else ""
        if not titulo or len(titulo) < 8:
            continue

        # Hora: El Espectador publica en UTC → restar 5 h para obtener COT
        dt_tag = card.select_one(".Card-Datetime")
        dt = None
        if dt_tag:
            raw = dt_tag.get_text(strip=True)
            mt = re.match(r"(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(a\.?\s*m\.?|p\.?\s*m\.?)?$", raw, re.I)
            if mt:
                hh = int(mt.group(1))
                mn = int(mt.group(2))
                ss = int(mt.group(3) or 0)
                mer = (mt.group(4) or "").lower().replace(" ", "").replace(".", "")
                if mer == "pm" and hh < 12:
                    hh += 12
                elif mer == "am" and hh == 12:
                    hh = 0
                today = datetime.now().date()
                try:
                    dt = datetime(today.year, today.month, today.day, hh, mn, ss) - timedelta(hours=5)
                except ValueError:
                    pass

        results.append(_item(titulo, href, medio, dt=dt))

    return results


# ─────────────────────────────────────────────────────────────────────────────
# 2. El Colombiano
#    URL: https://www.elcolombiano.com/
#    Items: article[iteridart]
#    URL real: div[data-urldestination]
#    Título: div.div_iter_title
#    Fecha: span.hora-noticia → "hace X horas/minutos"
# ─────────────────────────────────────────────────────────────────────────────
def scrape_elcolombiano() -> List[Dict]:
    medio = "El Colombiano"
    base  = "https://www.elcolombiano.com"
    soup  = _soup(base)
    if not soup:
        return []

    results, seen = [], set()
    for art in soup.find_all("article", attrs={"iteridart": True}):
        url_div = art.select_one("[data-urldestination]")
        if not url_div:
            continue
        href = _abs_url(url_div.get("data-urldestination", ""), base)
        if not href or href in seen or href.rstrip("/") == base:
            continue
        seen.add(href)

        title_div = art.select_one(".div_iter_title")
        titulo = title_div.get_text(strip=True) if title_div else ""
        if not titulo or len(titulo) < 8:
            continue

        hora_tag = art.select_one(".hora-noticia")
        dt = _parse_dt(hora_tag.get_text(strip=True)) if hora_tag else None

        results.append(_item(titulo, href, medio, dt=dt))

    return results[:30]


# ─────────────────────────────────────────────────────────────────────────────
# 3. El País
#    URL: https://elpais.com/ultimas-noticias/
#    Items: article.c  — el link del artículo contiene una fecha en la URL
#    Fecha: time[datetime]
# ─────────────────────────────────────────────────────────────────────────────
def scrape_elpais() -> List[Dict]:
    medio = "El País"
    base  = "https://elpais.com"
    soup  = _soup(f"{base}/ultimas-noticias/")
    if not soup:
        return []

    results, seen = [], set()
    for art in soup.find_all("article"):
        h = art.select_one("h2, h3, h4")
        if not h:
            continue
        titulo = h.get_text(strip=True)
        if not titulo or len(titulo) < 8:
            continue

        # El enlace del artículo contiene la fecha: /seccion/YYYY-MM-DD/titulo-...
        href = ""
        for a in art.find_all("a", href=True):
            val = a.get("href", "")
            if re.search(r"/\d{4}-\d{2}-\d{2}/", val):
                href = _abs_url(val, base)
                break
        if not href or href in seen:
            continue
        seen.add(href)

        time_tag = art.select_one("time[datetime]")
        dt = _dt_from_time_tag(time_tag)
        results.append(_item(titulo, href, medio, dt=dt))

    return results[:30]


# ─────────────────────────────────────────────────────────────────────────────
# 4. Vanguardia
#    URL: https://www.vanguardia.com/colombia/
#    Items: div con clase vg-storyCardBase (styled-components)
#    Fecha: span.date-container → "HH:MM a. m./p. m."
# ─────────────────────────────────────────────────────────────────────────────
def scrape_vanguardia() -> List[Dict]:
    medio = "Vanguardia"
    base  = "https://www.vanguardia.com"
    soup  = _soup(f"{base}/colombia/")
    if not soup:
        return []

    results, seen = [], set()
    for card in soup.find_all(class_=re.compile(r"vg-storyCardBase__StoryCardBase")):
        a = card.select_one("a[href]")
        if not a:
            continue
        href = _abs_url(a.get("href", ""), base)
        if not href or href in seen or href.rstrip("/") == base:
            continue
        seen.add(href)

        titulo_tag = card.select_one("h2, h3, h4")
        titulo = titulo_tag.get_text(strip=True) if titulo_tag else a.get_text(strip=True)
        if not titulo or len(titulo) < 8:
            continue

        date_span = card.select_one(".date-container")
        dt = _parse_dt(date_span.get_text(strip=True)) if date_span else None

        results.append(_item(titulo, href, medio, dt=dt))

    return results[:30]


# ─────────────────────────────────────────────────────────────────────────────
# 5. El Universal (Cartagena)
#    URL: https://www.eluniversal.com.co/
#    Items: div con clase eu-storyCardBase (styled-components)
# ─────────────────────────────────────────────────────────────────────────────
def scrape_eluniversal() -> List[Dict]:
    medio = "El Universal"
    base  = "https://www.eluniversal.com.co"
    soup  = _soup(base)
    if not soup:
        return []

    results, seen = [], set()
    for card in soup.find_all(class_=re.compile(r"eu-storyCardBase__StoryCardBase")):
        a = card.select_one("a[href]")
        if not a:
            continue
        href = _abs_url(a.get("href", ""), base)
        if not href or href in seen or href.rstrip("/") == base:
            continue
        seen.add(href)

        titulo_tag = card.select_one("h2, h3, h4")
        titulo = titulo_tag.get_text(strip=True) if titulo_tag else a.get_text(strip=True)
        if not titulo or len(titulo) < 8:
            continue

        # Intentar fecha en atributo o span
        time_tag = card.select_one("time[datetime]")
        dt = _dt_from_time_tag(time_tag)
        if not dt:
            for sel in (".date-container", "[class*='date']", "[class*='fecha']"):
                tag = card.select_one(sel)
                if tag:
                    dt = _parse_dt(tag.get_text(strip=True))
                    if dt:
                        break

        results.append(_item(titulo, href, medio, dt=dt))

    return results[:25]


# ─────────────────────────────────────────────────────────────────────────────
# 6. Noticias Caracol
#    URL: https://www.noticiascaracol.com/noticias
#    Fuente: JSON-LD NewsArticle (30 artículos con datePublished en COT -05:00)
#    Los timestamps del DOM son renderizados por JS y no están en HTML estático.
# ─────────────────────────────────────────────────────────────────────────────
def scrape_noticiascaracol() -> List[Dict]:
    import json as _json
    medio = "Noticias Caracol"
    base  = "https://www.noticiascaracol.com"
    r     = _get(f"{base}/noticias")
    if not r:
        return []

    soup = BeautifulSoup(r.content, "html.parser")
    results, seen = [], set()

    hoy_cot = (datetime.utcnow() - timedelta(hours=5)).date()

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = _json.loads(script.string or "")
            if not isinstance(data, dict) or data.get("@type") != "NewsArticle":
                continue
            titulo = (data.get("headline") or "").strip()
            url = (data.get("url") or
                   (data.get("mainEntityOfPage") or {}).get("@id") or "")
            fecha_str = data.get("datePublished", "")
            if not titulo or len(titulo) < 8 or not url or url in seen:
                continue
            dt = _parse_dt(fecha_str)
            if not dt or dt.date() != hoy_cot:   # solo artículos de hoy en COT
                continue
            seen.add(url)
            results.append(_item(titulo, url, medio, dt=dt))
        except Exception:
            pass

    return results[:30]


# ─────────────────────────────────────────────────────────────────────────────
# 7. Noticias RCN
#    URL: https://www.noticiasrcn.com/colombia
#    Items: div.news  →  div.post-h  →  a[href] con título en h2/h3
#    Fecha: no visible en HTML estático
# ─────────────────────────────────────────────────────────────────────────────
def scrape_noticiasrcn() -> List[Dict]:
    medio = "Noticias RCN"
    base  = "https://www.noticiasrcn.com"
    soup  = _soup(f"{base}/colombia")
    if not soup:
        return []

    results, seen = [], set()
    for post in soup.find_all("div", class_=re.compile(r"^post-h$")):
        a = post.select_one("a[href]")
        if not a:
            continue
        href = _abs_url(a.get("href", ""), base)
        if not href or href in seen:
            continue
        seen.add(href)

        titulo_tag = post.select_one("h2, h3, h4")
        titulo = titulo_tag.get_text(strip=True) if titulo_tag else a.get_text(strip=True)
        if not titulo or len(titulo) < 8:
            continue

        # post-time :created = Unix timestamp en ms (COT = UTC-5)
        post_time = post.select_one("post-time[\\:created]")
        dt = None
        if post_time:
            try:
                ts_ms = int(post_time.get(":created", 0))
                dt = datetime.utcfromtimestamp(ts_ms / 1000) - timedelta(hours=5)
            except (ValueError, TypeError):
                pass
        if not dt:
            time_tag = post.select_one("time[datetime]")
            dt = _dt_from_time_tag(time_tag)

        results.append(_item(titulo, href, medio, dt=dt))

    return results[:25]


# ─────────────────────────────────────────────────────────────────────────────
# 8. Teleantioquia
#    URL: https://www.teleantioquia.co/noticias
#    Nota: página renderizada por JS — sin contenido estático. Se retorna vacío.
# ─────────────────────────────────────────────────────────────────────────────
def scrape_teleantioquia() -> List[Dict]:
    print("  [i] Teleantioquia: renderizado por JS, sin resultados estáticos.")
    return []


# ─────────────────────────────────────────────────────────────────────────────
# 9. Canal Trece
#    URL: RSS https://canaltrece.com.co/feed/
#    Estructura: RSS con pubDate RFC 2822
# ─────────────────────────────────────────────────────────────────────────────
def scrape_canaltrece() -> List[Dict]:
    medio = "Canal Trece"
    base  = "https://canaltrece.com.co"
    r = _get(f"{base}/feed/")
    if not r:
        return []

    rss = BeautifulSoup(r.content, "lxml-xml")
    results = []
    for it in rss.find_all("item"):
        title_tag = it.find("title")
        link_tag  = it.find("link")
        pub_tag   = it.find("pubDate")
        if not title_tag:
            continue
        titulo = title_tag.get_text(strip=True)
        # En lxml-xml <link> puede ser vacío; URL en next_sibling
        if link_tag:
            href = link_tag.get_text(strip=True) or str(link_tag.next_sibling or "").strip()
        else:
            href = ""
        fecha_str = pub_tag.get_text(strip=True) if pub_tag else ""
        dt = _parse_dt(fecha_str)
        results.append(_item(titulo, href, medio, dt=dt))

    return results[:30]


# ─────────────────────────────────────────────────────────────────────────────
# 10. Blu Radio
#     URL: https://www.bluradio.com/noticias
#     Items: div.List-items-item  →  a[aria-label][href]  (mismo CMS que Caracol)
# ─────────────────────────────────────────────────────────────────────────────
def scrape_bluradio() -> List[Dict]:
    medio = "Blu Radio"
    base  = "https://www.bluradio.com"
    soup  = _soup(f"{base}/noticias")
    if not soup:
        return []

    results, seen = [], set()
    for item in soup.find_all("div", class_=re.compile(r"List-items-item")):
        a = item.select_one("a[aria-label][href]")
        if not a:
            a = item.select_one("a[href]")
        if not a:
            continue
        href = _abs_url(a.get("href", ""), base)
        if not href or href in seen:
            continue
        # Filtrar show-links de radio (URL de solo un nivel: /nombre-programa)
        path_parts = [p for p in href.replace(base, "").split("/") if p]
        if len(path_parts) < 2:
            seen.add(href)
            continue
        seen.add(href)

        titulo = a.get("aria-label", "") or a.get("title", "") or a.get_text(strip=True)
        if not titulo or len(titulo) < 10:
            continue

        dt = None
        promo_ts = item.select_one("[data-timestamp]")
        if promo_ts:
            try:
                ts_ms = int(promo_ts.get("data-timestamp", 0))
                dt = datetime.utcfromtimestamp(ts_ms / 1000) - timedelta(hours=5)
            except (ValueError, TypeError):
                pass
        if not dt:
            time_tag = item.select_one("time[datetime]")
            dt = _dt_from_time_tag(time_tag)
        if not dt:
            ts_tag = item.select_one("[class*='Promo-timestamp']")
            if ts_tag:
                dt = _parse_relative(ts_tag.get_text(strip=True))

        results.append(_item(titulo, href, medio, dt=dt))

    return results[:30]


# ─────────────────────────────────────────────────────────────────────────────
# 11. Semana
#     URL: https://www.semana.com/
#     Items: article  →  a[href], h2/h3 (título)
#     Fecha: span.text-neutral-500 → "hace 3 horas" / "hace 57 min"
# ─────────────────────────────────────────────────────────────────────────────
def scrape_semana() -> List[Dict]:
    medio = "Semana"
    base  = "https://www.semana.com"
    soup  = _soup(base)
    if not soup:
        return []

    results, seen = [], set()
    for art in soup.find_all("article"):
        a = art.select_one("a[href]")
        if not a:
            continue
        href = _abs_url(a.get("href", ""), base)
        if not href or href in seen:
            continue
        seen.add(href)

        titulo_tag = art.select_one("h2, h3, h4")
        titulo = titulo_tag.get_text(strip=True) if titulo_tag else a.get_text(strip=True)
        if not titulo or len(titulo) < 8:
            continue

        # Semana muestra "hace X horas/min" en un span con clase text-neutral-500
        dt = None
        time_tag = art.select_one("time[datetime]")
        if time_tag:
            dt = _dt_from_time_tag(time_tag)
        if not dt:
            for span in art.find_all("span"):
                txt = span.get_text(strip=True)
                if re.match(r"hace\s+", txt, re.I):
                    dt = _parse_relative(txt)
                    if dt:
                        break

        results.append(_item(titulo, href, medio, dt=dt))

    return results[:30]


# ─────────────────────────────────────────────────────────────────────────────
# 12. La Silla Vacía
#     URL: https://www.lasillavacia.com/
#     Contenedor: wp-block-column.home-en-vivo-desktop
#     Items: time.entry-date.published[datetime]  (WordPress)
# ─────────────────────────────────────────────────────────────────────────────
def scrape_lasillavacia() -> List[Dict]:
    medio = "La Silla Vacía"
    base  = "https://www.lasillavacia.com"
    soup  = _soup(base)
    if not soup:
        return []

    # Buscar el contenedor live
    container = None
    for tag in soup.find_all(class_=True):
        if "home-en-vivo-desktop" in " ".join(tag.get("class", [])):
            container = tag
            break
    if not container:
        # Fallback: buscar por home-en-vivo
        for tag in soup.find_all(class_=True):
            if "home-en-vivo" in " ".join(tag.get("class", [])):
                container = tag
                break
    if not container:
        container = soup

    results, seen = [], set()
    # Cada noticia tiene un <a href> directo a la noticia (no la categoría)
    for a in container.find_all("a", href=True):
        href = _abs_url(a.get("href", ""), base)
        # Filtrar: los artículos tienen rutas tipo /silla-llena/... o /en-vivo/...
        if not href or href in seen or href.rstrip("/") == base:
            continue
        # Ignorar links de categorías que son solo /categoria/
        if re.match(r"https://www\.lasillavacia\.com/category/", href):
            continue
        seen.add(href)

        # Buscar heading en el contexto inmediato del link
        titulo_tag = a.select_one("h2, h3, h4") or a.find_parent(
            lambda t: t.name in ("article", "div", "li") and t.select_one("h2,h3,h4")
        )
        if titulo_tag and hasattr(titulo_tag, "select_one"):
            titulo_tag = titulo_tag.select_one("h2, h3, h4")
        titulo = titulo_tag.get_text(strip=True) if titulo_tag else a.get_text(strip=True)
        if not titulo or len(titulo) < 8:
            continue

        # Buscar time[datetime] en el contenedor padre
        parent = a.find_parent(["article", "div", "li"])
        time_tag = parent.select_one("time[datetime]") if parent else None
        dt = _dt_from_time_tag(time_tag)

        results.append(_item(titulo, href, medio, dt=dt))

    return results[:20]


# ─────────────────────────────────────────────────────────────────────────────
# 13. Cambio Colombia
#     URL: https://cambiocolombia.com/
#     Contenedor: div.bg-[#F5F5F5] > div.relative > div.overflow-hidden >
#                 div.flex > N slides (min-w-0 shrink-0 grow-0 pl-2 …)
#     Cada slide tiene exactamente un <a><h3> y un <time[datetime]>
# ─────────────────────────────────────────────────────────────────────────────
def scrape_cambiocolombia() -> List[Dict]:
    medio = "Cambio Colombia"
    base  = "https://cambiocolombia.com"
    soup  = _soup(base)
    if not soup:
        return []

    # Localizar el contenedor Al Día
    container = None
    for tag in soup.find_all(True):
        if "bg-[#F5F5F5]" in " ".join(tag.get("class", [])):
            container = tag
            break

    results, seen = [], set()

    if container:
        # Navegar hasta el div flex que envuelve los slides
        flex = None
        overflow = container.find("div", class_=re.compile(r"overflow-hidden"))
        if overflow:
            flex = overflow.find("div")  # el div.flex inmediato

        if flex:
            # Cada hijo directo del flex es un slide con 1 h3 y 1 time
            for slide in flex.find_all("div", recursive=False):
                a_tag = next((a for a in slide.find_all("a", href=True)
                              if a.find("h3")), None)
                if not a_tag:
                    continue
                href = _abs_url(a_tag.get("href", ""), base)
                if not href or href in seen or href.rstrip("/") == base:
                    continue
                seen.add(href)

                titulo = a_tag.find("h3").get_text(strip=True)
                if not titulo or len(titulo) < 8:
                    continue

                time_tag = slide.select_one("time[datetime]")
                dt = _dt_from_time_tag(time_tag)
                results.append(_item(titulo, href, medio, dt=dt))

    if not results:
        # Fallback genérico: h3 dentro de <a href>
        for h3 in soup.find_all("h3"):
            parent_a = h3.find_parent("a")
            if not parent_a:
                continue
            href = _abs_url(parent_a.get("href", ""), base)
            if not href or href in seen or href.rstrip("/") == base:
                continue
            seen.add(href)
            titulo = h3.get_text(strip=True)
            if not titulo or len(titulo) < 8:
                continue
            block = parent_a.find_parent(["div", "li", "article"])
            time_tag = block.select_one("time[datetime]") if block else None
            dt = _dt_from_time_tag(time_tag)
            results.append(_item(titulo, href, medio, dt=dt))

    return results[:25]


# ─────────────────────────────────────────────────────────────────────────────
# Registro y ejecución paralela
# ─────────────────────────────────────────────────────────────────────────────
SCRAPERS: List[tuple] = [
    ("ElEspectador",    scrape_elespectador),
    ("ElColombiano",    scrape_elcolombiano),
    ("ElPaís",          scrape_elpais),
    ("Vanguardia",      scrape_vanguardia),
    ("ElUniversal",     scrape_eluniversal),
    ("NoticiasCaracol", scrape_noticiascaracol),
    ("NoticiasRCN",     scrape_noticiasrcn),
    ("Teleantioquia",   scrape_teleantioquia),
    ("CanalTrece",      scrape_canaltrece),
    ("BluRadio",        scrape_bluradio),
    ("Semana",          scrape_semana),
    ("LaSillaVacía",    scrape_lasillavacia),
    ("CambioColombia",  scrape_cambiocolombia),
]


def run_all(max_workers: int = 6) -> List[Dict]:
    """Ejecuta todos los scrapers en paralelo y retorna noticias ordenadas de más reciente a más antigua."""
    all_news: List[Dict] = []

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(fn): name for name, fn in SCRAPERS}
        for future in as_completed(futures):
            name = futures[future]
            try:
                items = future.result()
                print(f"  [✓] {name}: {len(items)} noticias")
                all_news.extend(items)
            except Exception as e:
                print(f"  [✗] {name}: {e}")

    # Ordenar: más reciente primero; items sin fecha (datetime.min) al final
    all_news.sort(key=lambda x: x.get("_dt", datetime.min), reverse=True)

    for item in all_news:
        item.pop("_dt", None)

    return all_news


def to_excel(noticias: List[Dict], path: str = "ultimahora_medios.xlsx") -> str:
    """Exporta la lista de noticias a un archivo Excel con formato."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(noticias, columns=["fecha", "medio", "titulo", "url"])
    df.columns = ["Fecha", "Medio", "Titular", "URL"]
    df.to_excel(path, index=False, engine="openpyxl")

    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Últimas Noticias"

    # Encabezados
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Cebra de filas y link clickeable
    fill_alt = PatternFill("solid", fgColor="EEF3F8")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    link_font = Font(color="0563C1", underline="single")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = fill_alt if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.border = border
            if fill.fill_type:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 3))

        # Columna URL → hipervínculo
        url_cell = ws.cell(row=row_idx, column=4)
        url_val = url_cell.value or ""
        if url_val.startswith("http"):
            url_cell.hyperlink = url_val
            url_cell.font = link_font
            url_cell.value = url_val

    # Anchos de columna
    ws.column_dimensions["A"].width = 22   # Fecha
    ws.column_dimensions["B"].width = 22   # Medio
    ws.column_dimensions["C"].width = 70   # Titular
    ws.column_dimensions["D"].width = 60   # URL

    # Altura de encabezado
    ws.row_dimensions[1].height = 22

    # Freeze panes
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return path


if __name__ == "__main__":
    import os

    print(f"Scraping {len(SCRAPERS)} medios...\n")
    noticias = run_all()
    print(f"\nTotal: {len(noticias)} noticias\n")

    # Generar Excel
    archivo = to_excel(noticias)
    print(f"Excel guardado en: {os.path.abspath(archivo)}")

    # Vista previa en consola
    print(f"\n{'FECHA':<22}  {'MEDIO':<20}  TITULAR")
    print("-" * 95)
    for n in noticias[:20]:
        print(f"{n['fecha']:<22}  {n['medio']:<20}  {n['titulo'][:55]}")
        print(f"  → {n['url']}")
