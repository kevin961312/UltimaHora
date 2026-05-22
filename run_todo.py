"""
run_todo.py — Ejecuta TODOS los scrapers (entidades oficiales + medios de comunicación)
y genera un único Excel con toda la información del día.

Columnas: Fecha | Tipo | Fuente | Titular | URL
"""

import os
import sys
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict

from scrapers import SCRAPERS as GOV_SCRAPERS
from Scrapermedios import SCRAPERS as MEDIA_SCRAPERS, _parse_dt


# ─────────────────────────────────────────────────────────────────────────────
# Logger — duplica stdout al archivo de log
# ─────────────────────────────────────────────────────────────────────────────

class _Tee:
    """Redirige stdout a terminal + archivo de log simultáneamente."""
    def __init__(self, file):
        self._file = file
        self._stdout = sys.stdout

    def write(self, data):
        self._stdout.write(data)
        self._file.write(data)

    def flush(self):
        self._stdout.flush()
        self._file.flush()

    def fileno(self):
        return self._stdout.fileno()


_log_file = None
LOG_PATH = "scrape_log.txt"


def _start_log():
    global _log_file
    _log_file = open(LOG_PATH, "w", encoding="utf-8")
    sys.stdout = _Tee(_log_file)


def _stop_log():
    global _log_file
    sys.stdout = sys.stdout._stdout  # restaurar stdout original
    if _log_file:
        _log_file.close()
        _log_file = None


# ─────────────────────────────────────────────────────────────────────────────
# Ejecución paralela
# ─────────────────────────────────────────────────────────────────────────────

def _run(scrapers, tipo_label: str, fuente_field: str, max_workers: int) -> List[Dict]:
    all_items: List[Dict] = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(fn): name for name, fn in scrapers}
        for future in as_completed(futures):
            name = futures[future]
            try:
                items = future.result()
                print(f"  [✓] {name}: {len(items)} noticias")
                for item in items:
                    dt = item.get("_dt") or _parse_dt(item.get("fecha", "")) or datetime.min
                    all_items.append({
                        "tipo":    tipo_label,
                        "fuente":  item[fuente_field],
                        "fecha":   item["fecha"],
                        "titulo":  item["titulo"],
                        "url":     item["url"],
                        "_dt":     dt,
                    })
            except Exception as e:
                print(f"  [✗] {name}: {e}")
    return all_items


def _block_sort(items: List[Dict]) -> List[Dict]:
    """Agrupa por fuente y ordena los bloques por la noticia más reciente de cada fuente."""
    from collections import defaultdict
    groups: dict = defaultdict(list)
    for item in items:
        groups[item["fuente"]].append(item)

    # Ordenar artículos dentro de cada bloque (más reciente primero)
    for fuente in groups:
        groups[fuente].sort(key=lambda x: x["_dt"], reverse=True)

    # Ordenar bloques por el artículo más reciente de cada fuente
    sorted_fuentes = sorted(groups.keys(), key=lambda f: groups[f][0]["_dt"], reverse=True)

    result = []
    for fuente in sorted_fuentes:
        result.extend(groups[fuente])
    return result


def run_all() -> List[Dict]:
    print(f"\n{'─'*60}")
    print(f"  ENTIDADES OFICIALES ({len(GOV_SCRAPERS)} scrapers)")
    print(f"{'─'*60}")
    gov   = _run(GOV_SCRAPERS,   "Entidad Oficial",       "ministerio", max_workers=8)

    print(f"\n{'─'*60}")
    print(f"  MEDIOS DE COMUNICACIÓN ({len(MEDIA_SCRAPERS)} scrapers)")
    print(f"{'─'*60}")
    media = _run(MEDIA_SCRAPERS, "Medio de Comunicación", "medio",      max_workers=6)

    combined = _block_sort(gov + media)
    for item in combined:
        item.pop("_dt", None)
    return combined


# ─────────────────────────────────────────────────────────────────────────────
# Exportar a Excel
# ─────────────────────────────────────────────────────────────────────────────

def to_excel(noticias: List[Dict], path: str) -> str:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(noticias, columns=["fecha", "tipo", "fuente", "titulo", "url"])
    df.columns = ["Fecha", "Tipo", "Fuente", "Titular", "URL"]
    df.to_excel(path, index=False, sheet_name="Todas las noticias", engine="openpyxl")

    wb = load_workbook(path)
    ws = wb.active

    # ── Encabezados ──────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Colores por tipo ─────────────────────────────────────────────────────
    fill_gov_even   = PatternFill("solid", fgColor="EAF2FB")  # azul muy suave
    fill_gov_odd    = PatternFill("solid", fgColor="F5FBFF")
    fill_media_even = PatternFill("solid", fgColor="FDFBE8")  # amarillo muy suave
    fill_media_odd  = PatternFill("solid", fgColor="FFFEF5")

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    link_font = Font(color="0563C1", underline="single", size=10)

    tipo_col = 2  # columna B = "Tipo"

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        tipo_val = ws.cell(row=row_idx, column=tipo_col).value or ""
        is_gov   = "Oficial" in tipo_val
        if is_gov:
            fill = fill_gov_even if row_idx % 2 == 0 else fill_gov_odd
        else:
            fill = fill_media_even if row_idx % 2 == 0 else fill_media_odd

        for cell in row:
            cell.border    = border
            cell.fill      = fill
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(cell.column == 4))  # col D = Titular

        # Columna URL → hipervínculo
        url_cell = ws.cell(row=row_idx, column=5)
        url_val  = url_cell.value or ""
        if url_val.startswith("http"):
            url_cell.hyperlink = url_val
            url_cell.font      = link_font
            url_cell.value     = url_val

    # ── Anchos ───────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22   # Fecha
    ws.column_dimensions["B"].width = 22   # Tipo
    ws.column_dimensions["C"].width = 30   # Fuente
    ws.column_dimensions["D"].width = 70   # Titular
    ws.column_dimensions["E"].width = 60   # URL

    ws.row_dimensions[1].height = 22

    # ── Freeze + autofilter ──────────────────────────────────────────────────
    ws.freeze_panes      = "A2"
    ws.auto_filter.ref   = ws.dimensions

    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# Punto de entrada
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import shutil

    archivo  = "ultimahora.xlsx"
    base_dir = os.path.dirname(os.path.abspath(__file__))

    _start_log()
    try:
        print(f"Iniciando scraping — {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        noticias = run_all()

        total_gov   = sum(1 for n in noticias if n["tipo"] == "Entidad Oficial")
        total_media = sum(1 for n in noticias if n["tipo"] == "Medio de Comunicación")

        print(f"\n{'═'*60}")
        print(f"  Entidades oficiales   : {total_gov:>4} noticias")
        print(f"  Medios de comunicación: {total_media:>4} noticias")
        print(f"  TOTAL                 : {len(noticias):>4} noticias")
        print(f"{'═'*60}\n")

        ruta = to_excel(noticias, archivo)
        print(f"Excel guardado en: {os.path.abspath(ruta)}")

        for destino in ("public", "docs"):
            dest_path = os.path.join(base_dir, destino, "ultimahora.xlsx")
            if os.path.isdir(os.path.dirname(dest_path)):
                shutil.copy2(archivo, dest_path)
                print(f"Copiado a:        {dest_path}")

        print(f"\nFin — {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    finally:
        _stop_log()

        # Copiar log a public/ y docs/ para que sea accesible en la página
        for destino in ("public", "docs"):
            dest_path = os.path.join(base_dir, destino, "scrape_log.txt")
            if os.path.isdir(os.path.dirname(dest_path)):
                shutil.copy2(LOG_PATH, dest_path)
                print(f"Log copiado a:    {dest_path}")
